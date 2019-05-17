import pandas as pd
import numpy as np
import os
import win32com.client
from openpyxl import load_workbook
from sklearn import svm
from sklearn.model_selection import StratifiedKFold
from sklearn.base import clone
from sklearn.metrics import confusion_matrix
from sklearn.metrics import accuracy_score
from sklearn.model_selection import cross_val_predict
from sklearn.svm import NuSVC
from sklearn.neural_network import MLPClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.ensemble import AdaBoostClassifier


def load_data_dane():
    #Wczytywanie danych
    #os.chdir('C:/Users/Krystian/Desktop/PracaInz')
    file = 'Dane.xlsx'
    pd.set_option('display.max_rows',10000)
    pd.set_option('display.width',10000)
    pets = pd.read_excel(file,sheet_name='Arkusz2')
    return pets


load = load_data_dane()
data_array = load.values


old_len = len(data_array)
#print(old_len)

def writer(arg_1, arg_2, arg_3, arg_4, arg_5, arg_6, arg_7, arg_8):

    df = pd.DataFrame([arg_1, arg_2, arg_3, arg_4, arg_5, arg_6, arg_7, arg_8])
    df = df.transpose()

    book = load_workbook('Dane.xlsx')
    writer = pd.ExcelWriter('Dane.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df.to_excel(writer, 'Arkusz3', startrow=1,index=False, header=False)
    writer.save()


writer("Pies","Pies","Średni",4.5,11,"Nie","Przyjazne","Neutralne")

number_elements = len(data_array)



#print(pd.DataFrame(data_array))
#print("********************* \n")



def function_normalize_naive(frame,value_old, value_new, index,size):
    if len(value_old) == 2 and len(value_new) == 2:
        for x in range(size):
            if frame[x][index] == value_old[0]:
                frame[x][index] = value_new[0]
            else:
                frame[x][index] = value_new[1]
    elif len(value_old) == 3 and len(value_new) == 3:
        for x in range(size):
            if frame[x][index] == value_old[0]:
                frame[x][index] = value_new[0]
            elif frame[x][index] == value_old[1]:
                frame[x][index] = value_new[1]
            else:
                frame[x][index] = value_new[2]
    elif len(value_old) == 4 and len(value_new) == 4:
        for x in range(size):
            if frame[x][index] == value_old[0]:
                frame[x][index] = value_new[0]
            elif frame[x][index] == value_old[1]:
                frame[x][index] = value_new[1]
            elif frame[x][index] == value_old[2]:
                frame[x][index] = value_new[2]
            else:
                frame[x][index] = value_new[3]




def function_min_max(frame,y,size):
    list = []

    for x in range(number_elements):
        list.append(data_array[x][y])

    max_value = max(list)
    min_value = min(list)


    for x in range(size):
        frame[x][y] = round((frame[x][y] - min_value) / (max_value - min_value), 2)


def init_function_min_max(pre,size):
    for x in range(9):
        function_min_max(pre,x,size)


def data_ark3():
    file = 'Dane.xlsx'
    data = pd.read_excel(file, sheet_name='Arkusz3')
    pre = data.values
    size = len(pre)

    # Gatunek
    index1 = 0
    y1 = [1, 2]
    d1 = ['Pies', 'Kot']
    function_normalize_naive(pre,d1, y1, index1, size)

    # Płec
    index2 = 1
    y2 = [1, 2, 3, 4]
    d2 = ['Pies', 'Suczka', 'Kocur', 'Kotka']
    function_normalize_naive(pre,d2, y2, index2, size)

    # Wielkosc
    index3 = 2
    y3 = [1, 2, 3]
    d3 = ['Mały', 'Średni', 'Duży']
    function_normalize_naive(pre,d3, y3, index3, size)

    # Sterylizacja
    index4 = 5
    y4 = [1, 2]
    d4 = ['Tak', 'Nie']
    function_normalize_naive(pre,d4, y4, index4, size)

    # Nastawienie do czlowieka
    index5 = 6
    y5 = [1, 2, 3]
    d5 = ['Przyjazne', 'Neutralne', 'Wrogie']
    function_normalize_naive(pre,d5, y5, index5, size)

    # Nastawienie do innych zwierzat
    index6 = 7
    y6 = [1, 2, 3]
    d6 = ['Przyjazne', 'Neutralne', 'Wrogie']
    function_normalize_naive(pre,d6, y6, index6, size)

    # Czy zaadoptowano
    index7 = 8
    y7 = [2, 1]
    d7 = ['Tak', 'Nie']
    function_normalize_naive(pre,d7, y7, index7, size)

    for x in range(size):
        pre[x][3] = int(pre[x][3])
        pre[x][4] = int(pre[x][4])

    # Normalizacja do zakresu [0,1]
    init_function_min_max(pre,size)


    #print(pd.DataFrame(pre))
    return pre



def deep_learning():
    #print(pd.DataFrame(data_array))

    #Gatunek
    index1 = 0
    y1 = [1, 2]
    d1 = ['Pies', 'Kot']
    function_normalize_naive(data_array,d1, y1, index1,number_elements)

    #Płec
    index2 = 1
    y2 = [1, 2, 3, 4]
    d2 = ['Pies', 'Suczka', 'Kocur', 'Kotka']
    function_normalize_naive(data_array,d2, y2, index2,number_elements)

    #Wielkosc
    index3 = 2
    y3 = [1, 2, 3]
    d3 = ['Mały', 'Średni', 'Duży']
    function_normalize_naive(data_array,d3, y3, index3,number_elements)

    #Sterylizacja
    index4 = 5
    y4 = [1, 2]
    d4 = ['Tak', 'Nie']
    function_normalize_naive(data_array,d4, y4, index4,number_elements)

    #Nastawienie do czlowieka
    index5 = 6
    y5 = [1, 2, 3]
    d5 = ['Przyjazne', 'Neutralne', 'Wrogie']
    function_normalize_naive(data_array,d5, y5, index5,number_elements)

    #Nastawienie do innych zwierzat
    index6 = 7
    y6 = [1, 2, 3]
    d6 = ['Przyjazne', 'Neutralne', 'Wrogie']
    function_normalize_naive(data_array,d6, y6, index6,number_elements)

    #Czy zaadoptowano
    index7 = 8
    y7 = [2, 1]
    d7 = ['Tak', 'Nie']
    function_normalize_naive(data_array,d7, y7, index7,number_elements)

    for x in range(number_elements):
        data_array[x][3] = int(data_array[x][3])
        data_array[x][4] = int(data_array[x][4])

    data_to_predict = data_ark3()



    #Normalizacja do zakresu [0,1]
    init_function_min_max(data_array,number_elements)



   #Cross validation + SVM
    crossX = np.delete(data_array,8,1)
    crossY = []
    for t1 in range(number_elements):
        crossY.append(data_array[t1][8])

    crossZ = np.delete(data_array,np.s_[0:8],1)
    results = list(map(int,crossY))

    cm_df = pd.DataFrame()
    acc_holder = []
    y_test_fold_holder = []
    y_pred_holder = []
    #Cs = [0.001, 0.01, 0.1, 1, 10]
    Cs = [25,50,75,100,200]
    #gammas = [0.001, 0.01, 0.1, 1, 10,100]
    gammas = [0.0001, 0.0005, 0.00001, 0.00005, 0.000001]
    func = ['linear','rbf']
    random_stat = 26



    gammas = [0.3]
    Cs_1 = [14,15]
    nu = [0.64]
    alpha_param = [0.1,0.5,0.01,0.05,0.001,0.005,0.0001,0.0005,0.00001,0.00005,0.000001,0.000005,0.0000001,0.0000005]
    solver_para = ['lbfgs']

    param_n_neigh = [5,10,15,20]
    param_weight = ['uniform','distance']
    param_algo = ['auto','ball_tree','kd_tree']

    param_n_esti = [50,100,150,200,250,300]

    for j in range(1):
        matrix_holder = np.ndarray([0])
        for i in  range(1):
            #clf = svm.SVC(kernel='rbf',gamma=gammas[i], C=Cs_1[j])
            clf = NuSVC(gamma=gammas[0],nu=nu[i])
            #clf = MLPClassifier(solver=solver_para[j],alpha=alpha_param[i],activation='logistic')
            #clf = KNeighborsClassifier(n_neighbors=1)
            #clf = AdaBoostClassifier(n_estimators=param_n_esti[i])
            matrix_holder = np.ndarray([0])
            skFolds = StratifiedKFold(n_splits=5,shuffle=True,random_state=random_stat)
            iter = 0

            #print("\nTest dla Ada  z parametrem n_estimators = ", 1)
            for train_index,test_index in skFolds.split(crossX,results):
                iter += 1
                clone_clf = clone(clf)
                X_train_folds = crossX[train_index]
                y_train_folds = (np.array(results)[train_index])
                X_test_fold = crossX[test_index]
                y_test_fold = (np.array(results)[test_index])

                clone_clf.fit(X_train_folds,y_train_folds)
                y_pred = clone_clf.predict(X_test_fold)



            #print("************************************************************************\n")
            y_train_pred = cross_val_predict(clf,crossX,results,cv=5)
            cmm = confusion_matrix(results,y_train_pred)
            matrix_score = np.append(matrix_holder,cmm)
            #print(matrix_score)

            #print("Macierz pomyłek")
            #print(cmm)


            acc_1 = (matrix_score[0] + matrix_score[3])/(matrix_score[0] + matrix_score[3] + matrix_score[1] + matrix_score[2])
            print("ACC1 = ", acc_1)



    #print(pd.DataFrame(data_array))




    score = clone_clf.predict(([
        [data_to_predict[0][0], data_to_predict[0][1], data_to_predict[0][2], data_to_predict[0][3], data_to_predict[0][4],
         data_to_predict[0][5], data_to_predict[0][6],data_to_predict[0][7]]]))
    #score = clone_clf.predict(([[0, 0.33, 0, 0.07, 0.17, 0, 0, 0.5]]))

    df = pd.DataFrame([score])
    df = df.transpose()

    book = load_workbook('Dane.xlsx')
    writer = pd.ExcelWriter('Dane.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df.to_excel(writer, 'Arkusz3', startrow=1, index=False, header=False,startcol=8)
    writer.save()

    return score





sc = deep_learning()

print(sc)

